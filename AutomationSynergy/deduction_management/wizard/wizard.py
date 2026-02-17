import openpyxl
from odoo import models, fields, api, _
from tempfile import TemporaryFile
from odoo.exceptions import warnings, UserError, ValidationError
import base64
from odoo import http
from odoo.http import request
import io
import xlsxwriter


class ContractImport(models.TransientModel):
    _name = "earnings.deduction.import"

    earnings_ded_excel_file = fields.Binary(string='Excel File')

    def import_earnings_ded_excel(self):
        if not self.earnings_ded_excel_file:
            raise ValidationError(_("Please Upload File to Import Earnings and Deduction !"))
        if self.earnings_ded_excel_file:
            file = base64.b64decode(self.earnings_ded_excel_file)
            earnings_ded_fileobj = TemporaryFile('wb+')
            earnings_ded_fileobj.write(file)
            earnings_ded_fileobj.seek(0)
            workbook = openpyxl.load_workbook(earnings_ded_fileobj, data_only=True)
            for sheet in workbook.worksheets:
                for row_cells in sheet.iter_rows(min_row=2):
                    earnings_ded_data = {}
                    employee = self.env['hr.employee'].search([('emp_id', '=', row_cells[0].value)])

                    if employee:
                        contract = self.env['hr.contract'].search([('employee_id', '=', employee.id), ('state', '=', 'open')])
                        # if not contract:
                        #     raise ValidationError(_("Employee with number  '%(emp_no)s' not having a running contract. ",
                        #                             emp_no=row_cells[0].value))
                        if contract :
                            rule = self.env['hr.salary.rule'].search([('code', '=', str(row_cells[2].value)),('is_earn_deduct','=',True)],limit=1)

                            if not rule:
                                raise ValidationError(_("Employee with number  '%(emp_no)s' not found matching salary rule. ",
                                                        emp_no=row_cells[0].value))


                            earnings_ded_data.update({
                                                  'employee_id': employee.id,
                                                    'contract_id':contract.id,
                                                  'department_id': employee.department_id.id,
                                                  'job_position': employee.job_id.id,
                                                  'type': str(row_cells[1].value),
                                                    'deduction_rule_id': rule.id if str(row_cells[1].value) == 'deductions' else False,
                                                    'allowance_rule_id': rule.id if str(row_cells[1].value) == 'earnings' else False,
                                                'deduction_amount': row_cells[3].value if row_cells[3].value else 0,
                                              'installment': row_cells[4].value if row_cells[4].value else 1,
                                              'payment_date': row_cells[5].value,
                                                'normal_ot_hours': row_cells[6].value if row_cells[6].value else 0,
                                                'public_ot_hours': row_cells[7].value if row_cells[7].value else 0,
                                                'friday_ot_hours': row_cells[8].value if row_cells[8].value else 0,
                                                'state':'validate'


                                                # 'company_id':employee.company_id.id

                                                  })
                            earnings_ded = self.env['hr.deduction'].create(earnings_ded_data)

                            earnings_ded.add_validators()
                            earnings_ded.compute_installment()

                    #     raise ValidationError(_("Employee with number  '%(emp_no)s' doesn't exist. ",
                    #                             emp_no=row_cells[0].value))
            # raise Warning(_("Excel Uploaded"))

    def export_employee_earnings(self):
        self.ensure_one()
        url = "/web/employee_earnings/template/%s" % (str(self.env.user.company_id.id))
        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'new',
        }


class EmployeesheetExport_Custom(http.Controller):

    @http.route("/web/employee_earnings/template/<int:company_id>", type="http", auth="user")
    def export_template(self, company_id=None, **kargs):
        employees = request.env['hr.employee'].sudo().search([('company_id', '=', int(company_id))])

        # XLSX
        excel = io.BytesIO()
        workbook = xlsxwriter.Workbook(excel, {'in_memory': True})
        worksheet = workbook.add_worksheet('Employee Profile')

        # Style
        head_style = workbook.add_format({'font_name': 'Arial', 'bold': True, 'font_size': 10, 'align': 'center'})
        head_style_holiday = workbook.add_format(
            {'font_name': 'Arial', 'bold': True, 'font_size': 10, 'align': 'center', 'bg_color': '#FF0000'})
        data_style_left = workbook.add_format({'font': 'Arial', 'font_size': 10, 'align': 'left'})
        data_style_right = workbook.add_format({'font': 'Arial', 'font_size': 10, 'align': 'right'})
        bck_data_style_right = workbook.add_format(
            {'font': 'Arial', 'font_size': 10, 'align': 'right', 'bg_color': '#D3D3D3'})
        data_style_center = workbook.add_format({'font': 'Arial', 'font_size': 10, 'align': 'center'})
        group_style_right = workbook.add_format(
            {'font': 'Arial', 'bold': True, 'font_size': 10, 'align': 'right', 'top': 1, })

        worksheet.set_column(0, 6, 10)
        worksheet.set_column(7, 37, 10)

        # Header
        worksheet.write(0, 0, 'Employee Number', head_style)
        worksheet.write(0, 1, 'Type', head_style)
        worksheet.write(0, 2, 'Rule', head_style)
        worksheet.write(0, 3, 'Amount', head_style)
        worksheet.write(0, 4, 'Installments', head_style)
        worksheet.write(0, 5, 'Payment Date', head_style)
        worksheet.write(0, 6, 'Normal OT', head_style)
        worksheet.write(0, 7, 'Public OT', head_style)
        worksheet.write(0, 8, 'Off Day OT', head_style)

        # worksheet.write(0, 6, 'Division', head_style)
        # worksheet.write(0, 7, 'GL Code', head_style)
        # worksheet.write(0, 8, 'Contract', head_style)
        # worksheet.write(0, 9, 'Supp/Debt', head_style)



        workbook.close()

        response = request.make_response(
            excel.getvalue(),
            headers=[
                ('Content-Type', 'application/octet-stream'),
                ('Content-Disposition', 'attachment; filename=EmployeeEarningDeduction.xlsx;')])

        return response

