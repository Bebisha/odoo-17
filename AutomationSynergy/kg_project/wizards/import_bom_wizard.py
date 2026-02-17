import base64
from io import BytesIO
import openpyxl
import tempfile
from odoo import models, fields, _
from odoo.exceptions import ValidationError
import logging
import xlsxwriter

_logger = logging.getLogger(__name__)

class ImportBOM(models.TransientModel):
    _name = "import.bom.wizard"
    _description = 'Importing the BOM'

    data_file = fields.Binary(string="Template File", required=True)
    filename = fields.Char(string="Filename", track_visibility="onchange", default='bom_template.xls')

    def export_file(self):
        wb = openpyxl.load_workbook(
            filename=BytesIO(base64.b64decode(self.data_file)),
            read_only=True
        )
        ws = wb.active

        # Setup variables
        product_quantities = {}
        bom_category = []
        description_position = None
        uom_position = None
        rating_position = None
        model_position = None

        for row_index, record in enumerate(ws.iter_rows(values_only=True), start=1):

            # HEADER: Detect all "Typical" columns (Row 1)
            if row_index == 1:
                for index, header in enumerate(record):
                    if header and isinstance(header, str) and 'typical' in header.lower():
                        category_name = self.env['bom.category'].search([('name', '=', header.strip())])
                        if not category_name:
                            raise ValidationError('This BOM Category is not Found - %s' % header)
                        bom_category.append((index, category_name))

            # COLUMN POSITIONS: Capture known headers (Row 2)
            elif row_index == 2:
                description_position = record.index('Item Description')
                uom_position = record.index('UOM')
                rating_position = record.index('Rating')
                model_position = record.index('Model No.')

            # DATA ROWS: Start actual product parsing (Row 3+)
            elif row_index >= 3:
                if record[0] and record[model_position] not in (None, 0):
                    # Find the product
                    products = self.env['product.product'].search([('default_code', '=', record[model_position])])
                    if not products:
                        continue
                    if len(products) > 1:
                        _logger.warning("Multiple products found with default_code '%s'. Using the first one.",
                                        record[model_position])
                    product_id = products[0]

                    # Find the UOM
                    uom_id = self.env['uom.uom'].search([('name', '=', record[uom_position])], limit=1)

                    for rec_index, bom_cat in bom_category:
                        qty = record[rec_index] or 0
                        unit_price = 0
                        subtotal = qty * unit_price

                        if (rec_index, bom_cat) not in product_quantities:
                            product_quantities[(rec_index, bom_cat)] = {}

                        product_cat = product_id.categ_id

                        if product_cat not in product_quantities[(rec_index, bom_cat)]:
                            product_quantities[(rec_index, bom_cat)][product_cat] = {}

                        if product_id not in product_quantities[(rec_index, bom_cat)][product_cat]:
                            product_quantities[(rec_index, bom_cat)][product_cat][product_id] = {
                                'qty': qty,
                                'uom': uom_id,
                                'unit_price': unit_price,
                                'subtotal': subtotal
                            }
                        else:
                            existing_product = product_quantities[(rec_index, bom_cat)][product_cat][product_id]
                            existing_product['qty'] += qty
                            existing_product['subtotal'] += subtotal
                            existing_product['unit_price'] = unit_price

        self.cleanup_empty_categories(product_quantities)

        if product_quantities:
            self.create_bom_master(product_quantities)

    def cleanup_empty_categories(self, product_quantities):
        boms_to_delete = []
        for bom in list(product_quantities.keys()):
            categories_to_delete = [
                cat_name for cat_name, products in product_quantities[bom].items()
                if all(product['qty'] == 0 for product in products.values())
            ]
            for cat_name in categories_to_delete:
                del product_quantities[bom][cat_name]

            if not product_quantities[bom]:
                boms_to_delete.append(bom)

        for bom in boms_to_delete:
            del product_quantities[bom]

    def create_bom_master(self, product_quantities):
        for bom_key, bom_value in product_quantities.items():
            bom_category = bom_key[1].id  # The bom.category record
            product_cat_keys = []
            vals = []
            for product_cat_key, product_cat_value in bom_value.items():
                vals.append((0, 0, {
                    'name': product_cat_key.name,
                    'display_type': 'line_section'
                }))
                for product_id, product_info in product_cat_value.items():
                    vals.append((0, 0, {
                        'product_id': product_id.id,
                        'display_type': False,
                        'qty': product_info['qty'],
                        'uom_id': product_info['uom'].id,
                        'unit_price': product_info['unit_price'],
                        'subtotal': product_info['subtotal']
                    }))
                product_cat_keys.append(product_cat_key.id)

            self.env['bom.master'].create({
                'show_bom_in': 'bom',
                'bom_category_id': bom_category,
                'category_ids': [(6, 0, product_cat_keys)],
                'bom_line_ids': vals
            })

    def auto_export_bom_file(self):
        # Create file in memory
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("BOM Export")

        # Formats
        bold = workbook.add_format({'bold': True})
        money = workbook.add_format({'num_format': '#,##0.00'})

        # Headers
        headers = [
            "BOM Category", "Product Category", "Document Code",
            "Product Name", "UOM", "Quantity", "Unit Price", "Subtotal"
        ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, bold)

        row = 1

        # Fetch all BOM records
        all_boms = self.env['bom.master'].search([])  # or filter by date or user if needed

        for bom in all_boms:
            bom_cat = bom.bom_category_id.name
            for line in bom.bom_line_ids:
                if line.display_type:
                    continue
                prod = line.product_id
                worksheet.write(row, 0, bom_cat)
                worksheet.write(row, 1, prod.categ_id.name if prod.categ_id else '')
                worksheet.write(row, 2, prod.default_code or '')
                worksheet.write(row, 3, prod.name or '')
                worksheet.write(row, 4, line.uom_id.name if line.uom_id else '')
                worksheet.write(row, 5, line.qty)
                worksheet.write(row, 6, line.unit_price, money)
                worksheet.write(row, 7, line.subtotal, money)
                row += 1

        workbook.close()
        output.seek(0)
        file_data = output.read()
        output.close()

        # Encode to base64
        b64_file = base64.b64encode(file_data)
        filename = "BOM_Export.xlsx"

        # Save as attachment
        self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': b64_file,
            'res_model': 'bom.master',
            'res_id': all_boms[0].id if all_boms else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })