import base64
import calendar
from odoo import fields, models, api
import xlsxwriter
import openpyxl
from io import BytesIO
from datetime import datetime, date

from odoo.exceptions import UserError

try:
    from base64 import encodebytes
except ImportError:
    from base64 import encodestring as encodebytes


class UpdateMilestoneWizard(models.Model):
    _name = 'update.milestone.wizard'

    type = fields.Selection(selection=[('sample', 'Download Sample Template'), ('import', 'Import Line')],
                            string="Update By",
                            required=True)
    file = fields.Binary(string="File", required=True)
    project_timeline_id = fields.Many2one('kg.project.timeline', string='Project Timeline')

    def action_download_milestones(self):
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError("No project timeline found in context.")

        project_timeline = self.env['kg.project.timeline'].browse(active_id)

        milestone_lines = project_timeline.milestone_line_ids
        if not milestone_lines:
            raise UserError("No milestone lines to export.")

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Milestones')

        headers = ['Name', 'Milestone']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        row = 1
        for line in milestone_lines:
            worksheet.write(row, 0, line.name or '')
            worksheet.write(row, 1, line.milestone_id.name or '')
            row += 1

        workbook.close()
        output.seek(0)
        xlsx_data = output.read()

        attachment = self.env['ir.attachment'].create({
            'name': f'{project_timeline.name}_milestones.xlsx',
            'type': 'binary',
            'datas': encodebytes(xlsx_data),
            'res_model': 'kg.project.timeline',
            'res_id': project_timeline.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        download_url = f'/web/content/{attachment.id}?download=true'
        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }

    def import_sample_template(self):
        if not self.file:
            raise UserError("Please upload a file to import.")

        # Decode and load the Excel file
        file_data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(filename=BytesIO(file_data), data_only=True)
        sheet = workbook.active

        # Get the active project timeline
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError("No project timeline found in context.")

        timeline = self.env['kg.project.timeline'].browse(active_id)
        project = timeline.name  # Get the linked project

        # Prepare mapping of existing milestones by name
        milestone_mapping = {line.name: line for line in timeline.milestone_line_ids}

        # Get all project milestones for lookup
        Milestone = self.env['project.milestone']
        existing_milestones = Milestone.search([('project_id', '=', project.id)])
        milestone_names = {m.name: m.id for m in existing_milestones}

        updated = 0
        created = 0
        milestones_created = 0

        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # skip header
            name = row[0].value
            milestone_name = row[1].value if len(row) > 1 else None

            if not name:
                continue  # skip empty rows

            # Determine if this should be a section (no milestone specified)
            is_section = not bool(milestone_name)

            # Prepare base values
            values = {
                'name': name,
                'project_id': timeline.id,
                'display_type': 'line_section' if is_section else False,
            }

            # Handle milestone (if not a section)
            milestone_id = False
            if not is_section:
                # Check if milestone exists
                if milestone_name not in milestone_names:
                    # Create new milestone
                    new_milestone = Milestone.create({
                        'name': milestone_name,
                        'project_id': project.id,
                    })
                    milestone_names[milestone_name] = new_milestone.id
                    milestones_created += 1

                milestone_id = milestone_names[milestone_name]
                values['milestone_id'] = milestone_id

            # Update existing line or create new one
            line = milestone_mapping.get(name)
            if line:
                line.write(values)
                updated += 1
            else:
                self.env['timeline.line'].create(values)
                created += 1

        message = f"""
        Successfully processed {created + updated} timeline lines:
        - {created} new lines created
        - {updated} existing lines updated
        - {milestones_created} new milestones created
        """

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import Completed',
                'message': message,
                'sticky': False,
                'next': {
                    'type': 'ir.actions.act_window_close',
                }
            }
        }