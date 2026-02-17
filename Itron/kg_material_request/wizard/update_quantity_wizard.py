import base64
import calendar
from odoo import fields, models, api
import xlsxwriter
import openpyxl
from io import BytesIO
from datetime import datetime, date

from odoo.exceptions import UserError

try:
    from base64 import encodebytes
except ImportError:
    from base64 import encodestring as encodebytes


class UpdateQuantityWizard(models.TransientModel):
    _name = 'update.quantity.wizard'

    type = fields.Selection(selection=[('sample', 'Download Sample Template'), ('import', 'Import Line')],
                            string="Update Qty",
                            required=True)
    file = fields.Binary(string="File", required=True)

    def download_sample_template(self):
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError("No stock picking found in context.")

        picking = self.env['stock.picking'].browse(active_id)

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Product Lines')
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 40)
        worksheet.set_column('C:C', 15)

        headers = ['Part Code', 'Product', 'Quantity', ]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        # Fill data from stock.move lines
        row = 1
        for move in picking.move_ids_without_package:
            worksheet.write(row, 0, move.product_id.default_code or '')  # Part Code
            worksheet.write(row, 1, move.product_id.display_name or '')  # Product Name
            worksheet.write(row, 2, '0')  # Quantity
            row += 1

        workbook.close()
        output.seek(0)
        xlsx_data = output.read()

        # Create an attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'{picking.name}_product_lines.xlsx',
            'type': 'binary',
            'datas': encodebytes(xlsx_data),
            'res_model': 'update.quantity.wizard',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        download_url = f'/web/content/{attachment.id}?download=true'
        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'self',
        }

    def import_sample_template(self):
        if not self.file:
            raise UserError("Please upload a file to import.")
        file_data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(filename=BytesIO(file_data), data_only=True)
        sheet = workbook.active
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError("No stock picking found in context.")

        picking = self.env['stock.picking'].browse(active_id)

        move_mapping = {move.product_id.display_name: move for move in picking.move_ids_without_package}
        move_mapping_by_code = {move.product_id.default_code: move for move in picking.move_ids_without_package if
                                move.product_id.default_code}

        updated = 0
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            part_code = row[0].value
            product_name = row[1].value
            quantity = row[2].value

            if not product_name and not part_code:
                continue

            # Try to find move by product name first, then by part code
            move = None
            if product_name:
                move = move_mapping.get(product_name)

            if not move and part_code:
                move = move_mapping_by_code.get(part_code)

            if move:
                try:
                    quantity = float(quantity)
                except (ValueError, TypeError):
                    raise UserError(f"Invalid quantity on row {i}: {quantity}")

                move.write({'qty': quantity,'is_verified': False})
                updated += 1
            else:
                identifier = product_name or part_code
                raise UserError(f"Product '{identifier}' not found in the picking move lines (row {i}).")

        return {
            'type': 'ir.actions.act_window_close',
        }