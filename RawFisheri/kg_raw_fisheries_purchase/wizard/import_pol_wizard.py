import base64
from tempfile import TemporaryFile

import openpyxl

from odoo import models, fields, _
from odoo.exceptions import UserError


class ImportPOLWizard(models.TransientModel):
    _name = "import.pol.wizard"
    _description = "Import PO Lines Wizard"

    name = fields.Char(string="Reference")
    is_po_import = fields.Boolean(default=False, string="Is PO Import")
    po_id = fields.Many2one("purchase.order", string="PO Ref")
    upload_file = fields.Binary(string="Upload Excel File", required=True)
    filename = fields.Char(string="File Name")

    def import_po(self):
        if self.upload_file:
            file_data = base64.b64decode(self.upload_file)
            fileobj = TemporaryFile('wb+')
            fileobj.write(file_data)
            fileobj.seek(0)
            workbook = openpyxl.load_workbook(fileobj, data_only=True)

            grouped_data = {}

            for sheet in workbook.worksheets:
                for row_cells in sheet.iter_rows(min_row=2):
                    row_num = row_cells[0].row
                    date = row_cells[1].value
                    vendor_name = row_cells[2].value
                    vessel_name = row_cells[3].value
                    product_code = row_cells[4].value
                    description = row_cells[5].value
                    qty = row_cells[6].value
                    uom_name = row_cells[7].value
                    unit_price = row_cells[8].value

                    if not vendor_name:
                        raise UserError(_("Row %s: Vendor is missing.") % row_num)

                    vendor = self.env['res.partner'].search([('name', '=', vendor_name)], limit=1)
                    if not vendor:
                        raise UserError(_("Row %s: Vendor '%s' not found.") % (row_num, vendor_name))

                    if not vessel_name:
                        raise UserError(_("Row %s: Vessel is missing.") % row_num)

                    vessel = self.env['sponsor.sponsor'].search([('name', '=', vessel_name)], limit=1)
                    if not vessel:
                        raise UserError(_("Row %s: Vessel '%s' not found.") % (row_num, vessel_name))

                    if not product_code:
                        raise UserError(_("Row %s: Part No is missing.") % row_num)

                    product = self.env['product.product'].search([('default_code', '=', product_code)], limit=1)
                    if not product:
                        raise UserError(_("Row %s: Product with code '%s' not found.") % (row_num, product_code))

                    if not uom_name:
                        raise UserError(_("Row %s: UOM is missing.") % row_num)

                    uom = self.env['uom.uom'].search([('name', '=', str(uom_name))], limit=1)
                    if not uom:
                        raise UserError(_("Row %s: UOM '%s' not found.") % (row_num, uom_name))

                    key = (date, vessel.id, vendor.id)
                    if key not in grouped_data:
                        grouped_data[key] = []

                    grouped_data[key].append({
                        'name': description,
                        'product_id': product.id,
                        'product_qty': qty,
                        'price_unit': unit_price,
                        'product_uom': uom.id,
                    })

            if grouped_data:
                for (date, vessel, vendor_id), lines in grouped_data.items():
                    po_vals = {
                        'partner_id': vendor_id,
                        'date_order': date,
                        'vessel_id': vessel,
                        'order_line': [],
                    }

                    po_line_vals = []
                    for line in lines:
                        po_line_vals.append((0, 0, {
                            'name': line['name'],
                            'product_id': line['product_id'],
                            'product_qty': line['product_qty'],
                            'price_unit': line['price_unit'],
                            'product_uom': line['product_uom'],
                        }))

                    po_vals['order_line'] = po_line_vals
                    self.env['purchase.order'].create(po_vals)

    def import_pol(self):
        if self.upload_file:
            file = base64.b64decode(self.upload_file)
            fileobj = TemporaryFile('wb+')
            fileobj.write(file)
            fileobj.seek(0)
            workbook = openpyxl.load_workbook(fileobj, data_only=True)

            for sheet in workbook.worksheets:
                for row_cells in sheet.iter_rows(min_row=2):
                    row_num = row_cells[0].row
                    product_code = row_cells[0].value
                    quantity = row_cells[1].value
                    price = row_cells[2].value
                    if not product_code:
                        raise UserError(_("Row %s: Product code is missing.") % row_num)

                    product = self.env['product.product'].search([('default_code', '=', product_code)], limit=1)
                    if not product:
                        raise UserError(_("Row %s: Product with code '%s' not found.") % (row_num, product_code))
                    else:
                        self.env['purchase.order.line'].create({
                            'order_id': self.po_id.id,
                            'product_id': product.id,
                            'product_qty': quantity,
                            'price_unit': price,
                            'name': product.name,
                            'product_uom': product.uom_po_id.id,
                        })

    def import_inventory_entries(self):
        if self.upload_file:
            file = base64.b64decode(self.upload_file)
            fileobj = TemporaryFile('wb+')
            fileobj.write(file)
            fileobj.seek(0)
            workbook = openpyxl.load_workbook(fileobj, data_only=True)

            for sheet in workbook.worksheets:
                for row_cells in sheet.iter_rows(min_row=2):
                    row_num = row_cells[0].row
                    # product_code = row_cells[0].value
                    warehouse_name = row_cells[0].value
                    location_name = row_cells[1].value

                    # if not product_code:
                    #     raise UserError(_("Row %s: Product code is missing.") % row_num)

                    if not warehouse_name:
                        raise UserError(_("Row %s: Warehouse is missing.") % row_num)

                    # product = self.env['product.product'].search([('default_code', '=', product_code)], limit=1)
                    # if not product:
                    #     raise UserError(_("Row %s: Product with code '%s' not found.") % (row_num, product_code))

                    warehouse = self.env['stock.warehouse'].search([('name', '=', warehouse_name)], limit=1)
                    if not warehouse:
                        raise UserError(_("Row %s: Warehouse '%s' not found.") % (row_num, warehouse_name))

                    location = self.env['res.country'].search([('name', '=', location_name)], limit=1)
                    if not location:
                        raise UserError(_("Row %s: Location '%s' not found.") % (row_num, location_name))

                    else:
                        self.env['inventory.update'].create({
                            # 'product_id': product.id,
                            'warehouse_id': warehouse.id,
                            'location_id': location.id,
                        })
